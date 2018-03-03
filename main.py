import glob, os
import time
import numpy
import multiprocessing
from datetime import datetime

from openpyxl import Workbook
from logger import logger
from tqdm import tqdm
import colorama

import face_recognition
from PIL import Image, ImageDraw
from scipy.spatial import distance


def parallel_run(file):
    (face_count, eye_area_list, eye_ear_list) = find_face(file)
    return (file, face_count, eye_area_list, eye_ear_list)


def main():
    start_time = time.time()
    logger.info('Begin')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = '檔案名'
    ws['B1'] = '人臉數目'
    ws['C1'] = '有眨眼人數'
    files = glob.glob("P:\\數碼相機\\2017-09-30\\*.jpg")
    # div = round(len(files) / 10)
    div = round(len(files) / 100)

    pool = multiprocessing.Pool(4)
    pbar = tqdm(total=len(files), unit="files")
    for i, (file, face_count, eye_area_list, eye_ear_list) in tqdm(enumerate(pool.imap(parallel_run, files))):
        has_eye_blink = []
        if len(eye_ear_list):
            eye_avg_ear_list = numpy.array(
                [(x[0] + x[1]) / 2 for x in eye_ear_list])
            has_eye_blink = eye_avg_ear_list[eye_avg_ear_list < 0.2]
        ws.append([os.path.basename(file), face_count] + [len(has_eye_blink)])
        pbar.update()
    pool.close()
    pool.join()

    # for i, file in tqdm(enumerate(files), total=len(files), unit="files"):
    #     (face_count, eye_area_list, eye_ear_list) = find_face(file)
    #     has_eye_blink = False
    #     if len(eye_ear_list):
    #         eye_avg_ear_list = numpy.array(
    #             [(x[0] + x[1]) / 2 for x in eye_ear_list])
    #         has_eye_blink = eye_avg_ear_list[eye_avg_ear_list < 0.2]
    #     ws.append([os.path.basename(file), face_count] + [len(has_eye_blink)])

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    wb.save("result" + datetime.now().strftime('%Y%m%d-%H%M%S') + ".xlsx")

    end_time = time.time()
    elapsed_time = end_time - start_time
    logger.info('End')
    logger.info('Elapsed %s', elapsed_time)


#https://stackoverflow.com/questions/25050311/extract-first-item-of-each-sublist-in-python
def eye_area(points):
    x = [point[0] for point in points]
    y = [point[1] for point in points]
    return PolyArea(x, y)


#https://www.pyimagesearch.com/2017/04/24/eye-blink-detection-opencv-python-dlib/
def eye_aspect_ratio(eye):
    # compute the euclidean distances between the two sets of
    # vertical eye landmarks (x, y)-coordinates
    A = distance.euclidean(eye[1], eye[5])
    B = distance.euclidean(eye[2], eye[4])

    # compute the euclidean distance between the horizontal
    # eye landmark (x, y)-coordinates
    C = distance.euclidean(eye[0], eye[3])

    # compute the eye aspect ratio
    ear = (A + B) / (2.0 * C)

    # return the eye aspect ratio
    return ear


def PolyArea(x, y):
    return 0.5 * numpy.abs(
        numpy.dot(x, numpy.roll(y, 1)) - numpy.dot(y, numpy.roll(x, 1)))


def find_face(filename):
    image = face_recognition.load_image_file(filename)
    face_locations = face_recognition.face_locations(image)
    # face_locations = face_recognition.face_locations(image, number_of_times_to_upsample=0, model="cnn")
    pass
    eye_area_list = []
    eye_ear_list = []
    if len(face_locations) > 0:
        face_landmarks_list = face_recognition.face_landmarks(
            image, face_locations)
        face_encodings = face_recognition.face_encodings(image, face_locations)
        pil_image = Image.fromarray(image)
        draw = ImageDraw.Draw(pil_image)

        for face_landmarks in face_landmarks_list:
            left_eye = face_landmarks['left_eye']
            right_eye = face_landmarks['right_eye']
            eye_area_list.append((eye_area(left_eye), eye_area(right_eye)))
            eye_ear_list.append((eye_aspect_ratio(left_eye),
                                 eye_aspect_ratio(right_eye)))

        # for (top, right, bottom, left), face_encoding in zip(
        #         face_locations, face_encodings):
        #     # See if the face is a match for the known face(s)
        #     # matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        #     # name = "Unknown"
        #     # If a match was found in known_face_encodings, just use the first one.
        #     # if True in matches:
        #     #     first_match_index = matches.index(True)
        #     #     name = known_face_names[first_match_index]

        #     # Draw a box around the face using the Pillow module
        #     #draw.rectangle(((left, top), (right, bottom)), outline=(0, 0, 255))
        #     # Draw a label with a name below the face
        #     #text_width, text_height = draw.textsize(name)
        #     #draw.text((left + 6, bottom - text_height - 5), name, fill=(255, 255, 255, 255))
        #     draw.rectangle(
        #         ((left + 10, top), (left, bottom)),
        #         fill=(255, 0, 0),
        #         outline=(255, 0, 0))
        #     draw.rectangle(
        #         ((left, top + 10), (right, top)),
        #         fill=(255, 0, 0),
        #         outline=(255, 0, 0))
        #     draw.rectangle(
        #         ((left, bottom - 10), (right, bottom)),
        #         fill=(255, 0, 0),
        #         outline=(255, 0, 0))
        #     draw.rectangle(
        #         ((right - 10, top), (right, bottom)),
        #         fill=(255, 0, 0),
        #         outline=(255, 0, 0))
        #     # draw.line(
        #     #     ((top, left), (top, right), (bottom, right), (bottom, left),
        #     #      (top, left)),
        #     #     fill=(255, 0, 0),
        #     #     width=15)

        # new_width = 800
        # new_height = new_width * pil_image.size[0] / pil_image.size[1]
        # # pil_image.thumbnail((new_width, new_height), Image.ANTIALIAS)
        # pil_image.save("r:\\" + os.path.basename(filename), 'JPEG', quality=60)

        # Remove the drawing library from memory as per the Pillow docs
        del draw
        del pil_image

    return (len(face_locations), eye_area_list, eye_ear_list)


main = main

if __name__ == "__main__":
    main()